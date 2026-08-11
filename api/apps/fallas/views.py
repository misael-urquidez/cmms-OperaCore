import csv
import io
import os

from django.conf import settings
from django.db import transaction
from django.http import HttpResponse
from django.core.exceptions import ValidationError as DjangoValidationError
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from django.shortcuts import get_object_or_404
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from rest_framework import generics, status
from rest_framework.response import Response
from rest_framework.views import APIView
from rest_framework.exceptions import ValidationError as DRFValidationError

from apps.usuarios.models import Trabajador
from apps.maquinaria.services import cambiar_estado_maquina
from . import models, serializers



class PingAPIView(APIView):

    def get(self, request):
        return Response({"modulo": "fallas", "status": "ok"}, status=status.HTTP_200_OK)


class TipoSeveridadListAPIView(generics.ListAPIView):

    queryset = models.TipoSeveridad.objects.all()
    serializer_class = serializers.TipoSeveridadSerializer


class TipoFallaListAPIView(generics.ListAPIView):

    queryset = models.TipoFalla.objects.all()
    serializer_class = serializers.TipoFallaSerializer


class TipoFallaCreateAPIView(generics.CreateAPIView):

    serializer_class = serializers.TipoFallaCreateSerializer


class TipoSeveridadCreateAPIView(generics.CreateAPIView):

    serializer_class = serializers.TipoSeveridadCreateSerializer


class TipoSeveridadDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    """GET trae un registro, PUT/PATCH lo edita, DELETE lo borra.
    Reutiliza el CreateSerializer para escritura porque tiene los mismos
    campos editables que el de detalle."""

    queryset = models.TipoSeveridad.objects.all()
    lookup_field = "codigo"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.TipoSeveridadCreateSerializer
        return serializers.TipoSeveridadDetailSerializer


class TipoFallaDetailAPIView(generics.RetrieveUpdateDestroyAPIView):

    queryset = models.TipoFalla.objects.all()
    lookup_field = "numeroRegistro"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.TipoFallaCreateSerializer
        return serializers.TipoFallaDetailSerializer


class MaquinaListAPIView(generics.ListAPIView):

    queryset = models.Maquina.objects.all()
    serializer_class = serializers.MaquinaSerializer


class EstadoReporteListAPIView(generics.ListAPIView):

    queryset = models.EstadoReporte.objects.all()
    serializer_class = serializers.EstadoReporteSerializer


class EstadoReporteCreateAPIView(generics.CreateAPIView):
    queryset = models.EstadoReporte.objects.all()
    serializer_class = serializers.EstadoReporteDetailSerializer


class EstadoReporteDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.EstadoReporte.objects.all()
    serializer_class = serializers.EstadoReporteDetailSerializer
    lookup_field = "codigo"


class ReporteFallaListAPIView(generics.ListAPIView):
    queryset = (
        models.ReporteFalla.objects
        .select_related("maquina", "trabajador", "tipo_severidad", "estado_reporte")
        .order_by("-fechaCreacion", "-horaCreacion")
    )
    serializer_class = serializers.ReporteFallaListSerializer

    def get_queryset(self):
        qs = (
            models.ReporteFalla.objects
            .select_related("maquina", "trabajador", "tipo_severidad")
            .order_by("-fechaCreacion", "-horaCreacion")
        )
        if self.request.query_params.get("trabajador"):
            qs = qs.filter(trabajador_id=self.request.query_params["trabajador"])
        if self.request.query_params.get("maquina"):
            qs = qs.filter(maquina_id=self.request.query_params["maquina"])
        return qs

#cambio

class ReporteFallaDetailAPIView(generics.RetrieveAPIView):

    queryset = models.ReporteFalla.objects.select_related(
        "maquina", "trabajador", "tipo_severidad"
    )
    serializer_class = serializers.ReporteFallaDetailSerializer


class ReporteFallaCreateAPIView(generics.CreateAPIView):

    serializer_class = serializers.ReporteFallaCreateSerializer

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        try:
            with transaction.atomic():
                reporte = serializer.save()
                cambiar_estado_maquina(
                    reporte.maquina_id, "FALLO", "reporte_falla", str(reporte.numeroRegistro),
                )
                data = serializers.ReporteFallaDetailSerializer(reporte).data
                return Response(data, status=status.HTTP_201_CREATED)
        except DjangoValidationError as e:
            raise DRFValidationError({"maquina": e.messages})


class ReporteFallaUpdateAPIView(generics.UpdateAPIView):

    queryset = models.ReporteFalla.objects.all()
    serializer_class = serializers.ReporteFallaUpdateSerializer
    def update(self, request, *args, **kwargs):
        response = super().update(request, *args, **kwargs)
        reporte = models.ReporteFalla.objects.get(pk=kwargs["pk"])

        tipo_falla_ids = request.data.getlist("tipo_falla_ids")
        if tipo_falla_ids:
            models.TipoReporte.objects.filter(reporte_falla=reporte).delete()
            for tf_id in tipo_falla_ids:
                models.TipoReporte.objects.create(
                    tipo_falla_id=int(tf_id),
                    reporte_falla=reporte,
                )

        imagen_file = request.FILES.get("imagen")
        if imagen_file:
            carpeta = os.path.join(settings.MEDIA_ROOT, "fallas")
            os.makedirs(carpeta, exist_ok=True)
            with open(os.path.join(carpeta, imagen_file.name), "wb+") as dest:
                for chunk in imagen_file.chunks():
                    dest.write(chunk)
            reporte.imagen = f"fallas/{imagen_file.name}"
            reporte.save(update_fields=["imagen"])

        return Response(
            serializers.ReporteFallaDetailSerializer(reporte).data,
            status=status.HTTP_200_OK,
        )
    
class TrabajadorListAPIView(generics.ListAPIView):
    """Listado ligero de técnicos activos para asignar o reasignar
    órdenes de mantenimiento."""

    queryset = Trabajador.objects.filter(
        actividad=True, rol__codigo="TECNI"
    ).order_by("nombre")
    serializer_class = serializers.TrabajadorLightSerializer


# ------------ TIPO_REPORTE (llave compuesta) ------------------------------
class TipoReporteListAPIView(generics.ListAPIView):
    queryset = models.TipoReporte.objects.all()
    serializer_class = serializers.ListTipoReporteSerializer


class TipoReporteCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTipoReporteSerializer


class TipoReporteDetailAPIView(generics.GenericAPIView):
    queryset = models.TipoReporte.objects.all()
    serializer_class = serializers.DetailTipoReporteSerializer

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.CreateTipoReporteSerializer
        return serializers.DetailTipoReporteSerializer

    def get_object(self):
        obj = generics.get_object_or_404(
            self.get_queryset(),
            tipo_falla=self.kwargs["tipo_falla"],
            reporte_falla=self.kwargs["reporte_falla"],
        )
        self.check_object_permissions(self.request, obj)
        return obj

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        serializer = self.get_serializer(obj)
        return Response(serializer.data)

    def put(self, request, *args, **kwargs):
        obj = self.get_object()
        serializer = self.get_serializer(obj, data=request.data)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def patch(self, request, *args, **kwargs):
        obj = self.get_object()
        serializer = self.get_serializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)

    def delete(self, request, *args, **kwargs):
        obj = self.get_object()
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)


class CatalogosReporteAPIView(APIView):
    """Junta los catalogos que usa el formulario de 'Reportar Falla' en
    una sola respuesta, para que el client no tenga que hacer N llamadas
    HTTP separadas y secuenciales cada vez que carga la pagina."""

    def get(self, request):
        maquinas_operativas = models.Maquina.objects.filter(estado_maquina="OPERA")

        data = {
            "severidades": serializers.TipoSeveridadSerializer(
                models.TipoSeveridad.objects.all(), many=True
            ).data,
            "tipos_falla": serializers.TipoFallaSerializer(
                models.TipoFalla.objects.all(), many=True
            ).data,
            "maquinas": serializers.MaquinaSerializer(maquinas_operativas, many=True).data,
            "estados": serializers.EstadoReporteSerializer(
                models.EstadoReporte.objects.all(), many=True
            ).data,
            "trabajadores": serializers.TrabajadorLightSerializer(
                models.Trabajador.objects.filter(actividad=True).order_by("nombre"),
                many=True,
            ).data,
        }
        return Response(data, status=status.HTTP_200_OK)


# ------------ EXPORTACIONES  -----------------------------------------
def _get_reporte_data(pk):
    qs = models.ReporteFalla.objects.select_related(
        "maquina", "trabajador", "tipo_severidad", "estado_reporte"
    )
    reporte = generics.get_object_or_404(qs, pk=pk)
    ser = serializers.ReporteFallaDetailSerializer(reporte)
    return ser.data


class ExportarReporteCSVAPIView(APIView):

    def get(self, request, pk):
        data = _get_reporte_data(pk)
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow(["Campo", "Valor"])
        writer.writerow(["# Reporte", data.get("numeroRegistro")])
        writer.writerow(["Asunto", data.get("asunto")])
        writer.writerow(["Maquina", data.get("maquina_nombre")])
        writer.writerow(["Trabajador", data.get("trabajador_nombre")])
        writer.writerow(["Severidad", data.get("tipo_severidad_nombre")])
        writer.writerow(["Estado", data.get("estado_reporte_nombre")])
        writer.writerow(["Fecha creacion", data.get("fechaCreacion")])
        writer.writerow(["Hora creacion", data.get("horaCreacion")])
        writer.writerow(["Fecha resolucion", data.get("fechaResolucion") or ""])
        writer.writerow(["Tiempo paro (hrs)", data.get("tiempoParo") or ""])
        writer.writerow(["Descripcion", data.get("descripcion") or ""])
        writer.writerow(["Causa raiz", data.get("causaRaiz") or ""])

        fallas = data.get("fallas_asociadas", [])
        nombres_falla = ", ".join(f["nombre"] for f in fallas) if fallas else ""
        writer.writerow(["Tipos de falla", nombres_falla])

        response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="reporte_falla_{pk}.csv"'
        return response


class ExportarReporteXLSXAPIView(APIView):

    def get(self, request, pk):
        data = _get_reporte_data(pk)
        wb = Workbook()
        ws = wb.active
        ws.title = "Reporte"

        header_font = Font(bold=True)
        labels = [
            "# Reporte", "Asunto", "Maquina", "Trabajador", "Severidad",
            "Estado", "Fecha creacion", "Hora creacion",
            "Fecha resolucion", "Tiempo paro (hrs)",
        ]
        values = [
            data.get("numeroRegistro"), data.get("asunto"),
            data.get("maquina_nombre"), data.get("trabajador_nombre"),
            data.get("tipo_severidad_nombre"), data.get("estado_reporte_nombre"),
            data.get("fechaCreacion"), data.get("horaCreacion"),
            data.get("fechaResolucion") or "", data.get("tiempoParo") or "",
        ]
        for row_idx, (label, value) in enumerate(zip(labels, values), 1):
            ws.cell(row=row_idx, column=1, value=label).font = header_font
            ws.cell(row=row_idx, column=2, value=str(value))

        desc_row = len(labels) + 2
        ws.cell(row=desc_row, column=1, value="Descripcion").font = header_font
        ws.cell(row=desc_row, column=2, value=data.get("descripcion") or "")
        ws.cell(row=desc_row + 1, column=1, value="Causa raiz").font = header_font
        ws.cell(row=desc_row + 1, column=2, value=data.get("causaRaiz") or "")

        fallas = data.get("fallas_asociadas", [])
        nombres_falla = ", ".join(f["nombre"] for f in fallas) if fallas else ""
        ws.cell(row=desc_row + 2, column=1, value="Tipos de falla").font = header_font
        ws.cell(row=desc_row + 2, column=2, value=nombres_falla)

        ws.column_dimensions["A"].width = 20
        ws.column_dimensions["B"].width = 60

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="reporte_falla_{pk}.xlsx"'
        return response


class ExportarReportePDFAPIView(APIView):

    def get(self, request, pk):
        data = _get_reporte_data(pk)
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []

        elements.append(Paragraph(f"Reporte de Falla #{data.get('numeroRegistro')}", styles["Title"]))
        elements.append(Spacer(1, 12))

        rows = [
            ["Campo", "Valor"],
            ["Asunto", data.get("asunto")],
            ["Maquina", data.get("maquina_nombre")],
            ["Trabajador", data.get("trabajador_nombre")],
            ["Severidad", data.get("tipo_severidad_nombre")],
            ["Estado", data.get("estado_reporte_nombre")],
            ["Fecha creacion", f"{data.get('fechaCreacion')} {data.get('horaCreacion') or ''}"],
            ["Fecha resolucion", data.get("fechaResolucion") or ""],
            ["Tiempo paro (hrs)", str(data.get("tiempoParo") or "")],
        ]

        table = Table(rows, colWidths=[140, 340])
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
            ("FONTSIZE", (0, 0), (-1, 0), 10),
            ("FONTNAME", (0, 1), (-1, -1), "Helvetica"),
            ("FONTSIZE", (0, 1), (-1, -1), 9),
            ("ALIGN", (0, 0), (-1, -1), "LEFT"),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.HexColor("#d1d5db")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.HexColor("#ffffff"), colors.HexColor("#f9fafb")]),
        ]))
        elements.append(table)
        elements.append(Spacer(1, 16))

        if data.get("descripcion"):
            elements.append(Paragraph("Descripcion", styles["Heading2"]))
            elements.append(Paragraph(data["descripcion"], styles["Normal"]))
            elements.append(Spacer(1, 12))

        if data.get("causaRaiz"):
            elements.append(Paragraph("Causa Raiz", styles["Heading2"]))
            elements.append(Paragraph(data["causaRaiz"], styles["Normal"]))
            elements.append(Spacer(1, 12))

        fallas = data.get("fallas_asociadas", [])
        if fallas:
            elements.append(Paragraph("Tipos de Falla Asociados", styles["Heading2"]))
            for f in fallas:
                elements.append(Paragraph(f"- {f['nombre']}", styles["Normal"]))

        doc.build(elements)
        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = f'attachment; filename="reporte_falla_{pk}.pdf"'
        return response



class ActualizarReporteFallaView(APIView):

    def post(self, request, pk):
        reporte = get_object_or_404(models.ReporteFalla, pk=pk)
        estado_anterior = reporte.estado_reporte

        # 1. Validación: Si el reporte estaba En Espera ('ENESP'), no permitir cambiar estado
        nuevo_estado = request.data.get("estado_reporte")
        if estado_anterior == "ENESP" and nuevo_estado != "ENESP":
            return Response(
                {
                    "error": "Un reporte en estado 'En Espera' no se puede editar."
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Usamos tu serializer existente para validar y actualizar el reporte
        serializer = serializers.ReporteFallaUpdateSerializer(
            reporte, data=request.data, partial=True
        )
        if serializer.is_valid():
            reporte_actualizado = serializer.save()
            maquina = reporte_actualizado.models.Maquina 

            # 2. Transición de estado de la Máquina:
            # Si cambia a Cancelado ('CANCE'), la máquina pasa a En Espera ('ENESP')
            if estado_anterior != "CANCE" and nuevo_estado == "CANCE":
                maquina.estado = "ENESP"  # Estado En Espera de la máquina
                maquina.save()

            # Si cambia de Cancelado ('CANCE') a Abierto ('ABIER'), la máquina vuelve a FALLO
            elif estado_anterior == "CANCE" and nuevo_estado == "ABIER":
                maquina.estado = "FALLO"  # Estado Fallo de la máquina
                maquina.save()

            return Response(serializer.data, status=status.HTTP_200_OK)

        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)