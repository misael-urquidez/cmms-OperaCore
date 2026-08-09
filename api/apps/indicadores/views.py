from django.db import connection
from django.db.utils import OperationalError
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import generics, status
from django.db.models import F
from apps.monitoreo.models import Indicador
from apps.mantenimiento.models import OrdenMantenimiento
from apps.inventario.models import Refaccion
from . import models
from . import serializers

import csv
import io
import re

from django.http import HttpResponse
from django.utils import timezone
import os

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from reportlab.lib.pagesizes import letter, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from svglib.svglib import svg2rlg
from reportlab.graphics import renderPDF
from apps.mantenimiento.views import BRAND_DARK, BRAND_ACCENT, BRAND_MUTED, BRAND_LINE, LOGO_PATH

PAGE_W_KPI, PAGE_H_KPI = landscape(letter)

LOGO_SVG_PATH = os.path.join(os.path.dirname(__file__), "assets", "logo.svg")
_logo_drawing = None
if os.path.isfile(LOGO_SVG_PATH):
    try:
        _logo_drawing = svg2rlg(LOGO_SVG_PATH)
    except Exception:
        _logo_drawing = None


# Whitelist de vistas de vistas_kpi.sql: el nombre real de la vista SIEMPRE
# sale de aqui, nunca del parametro crudo de la URL -- asi no hay forma de
# inyectar SQL aunque <vista> venga directo de un usuario.
VISTAS_KPI = {
    "estado-flota": "v_kpi_estado_flota",
    "reportes-atencion": "v_kpi_reportes_atencion",
    "stock": "v_kpi_stock",
    "fallas-por-maquina": "v_kpi_fallas_por_maquina",
    "top-fallas": "v_kpi_top_fallas",
    "horas-operacion": "v_kpi_horas_operacion",
    "mantenimiento-por-maquina": "v_kpi_mantenimiento_por_maquina",
    "indicadores-actuales": "v_kpi_indicadores_actuales",
    "disponibilidad-linea": "v_kpi_disponibilidad_linea",
    "monitoreo-predictivo": "v_kpi_monitoreo_predictivo",
}

TITULOS_KPI = {
    "estado-flota": "Estado de la flota",
    "indicadores-actuales": "Indicadores actuales",
    "disponibilidad-linea": "Disponibilidad por línea",
    "fallas-por-maquina": "Fallas por máquina",
    "top-fallas": "Top fallas",
    "mantenimiento-por-maquina": "Mantenimiento por máquina",
    "horas-operacion": "Horas de operación",
    "reportes-atencion": "Reportes en atención",
    "stock": "Stock de refacciones",
    "monitoreo-predictivo": "Monitoreo predictivo",
}

# Traduce alias crudos de SQL (CamelCase) a etiquetas legibles.
# Si una columna no está en el override, se separa automáticamente
# por mayúsculas: "TotalFallas" -> "Total Fallas".
COLUMNAS_OVERRIDE = {
    "MTTR": "MTTR (hrs)",
    "MTBF": "MTBF (hrs)",
    "Disponibilidad": "Disponibilidad (%)",
    "TotalFallas": "Total de fallas",
    "HorasOperacion": "Horas de operación",
    "FallasAbiertas": "Fallas abiertas",
    "OrdenesActivas": "Órdenes activas",
    "OrdenesEnProgreso": "Órdenes en progreso",
    "TipoFalla": "Tipo de falla",
    "StockMinimo": "Stock mínimo",
    "Refaccion": "Refacción",
    "Umbral": "Umbral de vibración",
    "Vibracion": "Vibración",
    "Excede": "¿Excede umbral?",
}


def _humanizar_columna(nombre):
    if nombre in COLUMNAS_OVERRIDE:
        return COLUMNAS_OVERRIDE[nombre]
    if nombre.isupper():
        return nombre
    return re.sub(r"(?<!^)(?=[A-Z])", " ", nombre)


def _formatear_valor(valor, columna=""):
    """Convierte valores crudos de la vista SQL en texto listo para
    mostrar: nulos -> '—', floats redondeados, fechas ISO -> dd/mm/aaaa,
    booleanos tipo 'Excede' -> Sí/No."""
    if valor is None or valor == "":
        return "—"
    if columna == "Excede":
        return "Sí" if valor else "No"
    if isinstance(valor, bool):
        return "Sí" if valor else "No"
    if isinstance(valor, float):
        texto = f"{valor:.1f}".rstrip("0").rstrip(".")
        if "disponibilidad" in columna.lower():
            texto += "%"
        return texto
    if isinstance(valor, int):
        return str(valor)
    if isinstance(valor, str):
        m = re.match(r"^(\d{4})-(\d{2})-(\d{2})(?:[T ](\d{2}):(\d{2}))?", valor)
        if m:
            y, mo, d, h, mi = m.groups()
            return f"{d}/{mo}/{y}" + (f" {h}:{mi}" if h else "")
        return valor
    return str(valor)


def _filas_a_dicts(cur):
    columnas = [c[0] for c in cur.description]
    data = [dict(zip(columnas, fila)) for fila in cur.fetchall()]
    for row in data:
        for k, v in row.items():
            if hasattr(v, "isoformat"):  # date/datetime -> string
                row[k] = v.isoformat()
    return data


class KPIVistaAPIView(APIView):
    """Endpoint generico de solo lectura para las 10 vistas de
    vistas_kpi.sql. GET /indicadores/v1/kpi/<vista>/ -> SELECT * FROM
    v_kpi_<vista correspondiente>. <vista> usa guiones (estado-flota),
    no el nombre real de la tabla."""

    def get(self, request, vista):
        tabla = VISTAS_KPI.get(vista)
        if tabla is None:
            return Response(
                {"detail": "Vista no encontrada.", "disponibles": list(VISTAS_KPI)},
                status=status.HTTP_404_NOT_FOUND,
            )
        with connection.cursor() as cur:
            cur.execute(f"SELECT * FROM {tabla}")
            data = _filas_a_dicts(cur)
        return Response(data, status=status.HTTP_200_OK)


def _fila_en_periodo(fila, fecha_inicio, fecha_fin):
    """Filtro generico de periodo: si la fila tiene columnas cuyo nombre
    contenga 'fecha' o 'periodo', verifica que el valor caiga dentro del
    rango [fecha_inicio, fecha_fin]. Si no hay columnas de fecha, la fila
    se considera siempre dentro del periodo."""
    if not fecha_inicio and not fecha_fin:
        return True

    for k, v in fila.items():
        if re.search(r"fecha|periodo", k, re.IGNORECASE):
            try:
                if isinstance(v, str):
                    v = timezone.datetime.fromisoformat(v).date()
                elif hasattr(v, "isoformat"):
                    v = v.date()
                else:
                    continue

                if fecha_inicio and v < fecha_inicio:
                    return False
                if fecha_fin and v > fecha_fin:
                    return False
            except (ValueError, TypeError):
                continue
    return True


def _generar_csv(vistas_data, nombre_archivo="reporte_kpi"):
    """Genera un CSV con una hoja por vista (separadas por una linea en blanco)."""
    output = io.StringIO()
    writer = csv.writer(output)

    writer.writerow(["OperaCore — Reporte de indicadores"])
    writer.writerow([f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')}"])
    writer.writerow([])

    for vista, data in vistas_data:
        if not data:
            continue

        writer.writerow([TITULOS_KPI.get(vista, vista)])

        headers_raw = list(data[0].keys())
        writer.writerow([_humanizar_columna(h) for h in headers_raw])

        for fila in data:
            writer.writerow([_formatear_valor(fila.get(h), h) for h in headers_raw])

        writer.writerow([])
        writer.writerow([])

    # BOM al inicio: evita que Excel muestre mal los acentos/ñ del CSV.
    contenido = "﻿" + output.getvalue()
    response = HttpResponse(contenido, content_type="text/csv")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.csv"'
    return response


def _generar_xlsx(vistas_data, nombre_archivo="reporte_kpi"):
    """Genera un XLSX con una hoja por vista: título, subtítulo, encabezado
    fijo con filtro, cebra en filas y columnas autoajustadas."""
    wb = Workbook()
    ws = wb.active
    wb.remove(ws)

    titulo_fill = PatternFill(start_color="12261E", end_color="12261E", fill_type="solid")
    titulo_font = Font(color="FFFFFF", bold=True, size=13)
    subtitulo_font = Font(color="6B7280", italic=True, size=9)
    header_fill = PatternFill(start_color="0369A1", end_color="0369A1", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    zebra_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    borde_fino = Border(bottom=Side(style="thin", color="E5E7EB"))

    for vista, data in vistas_data:
        if not data:
            continue

        ws = wb.create_sheet(title=TITULOS_KPI.get(vista, vista)[:31])
        headers_raw = list(data[0].keys())
        headers = [_humanizar_columna(h) for h in headers_raw]
        n_cols = len(headers)

        # Fila 1: título de marca
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=n_cols)
        celda_titulo = ws.cell(row=1, column=1, value=f"OperaCore — {TITULOS_KPI.get(vista, vista)}")
        celda_titulo.fill = titulo_fill
        celda_titulo.font = titulo_font
        celda_titulo.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[1].height = 24

        # Fila 2: subtítulo (fecha de generación + total de registros)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
        sub = f"Generado el {timezone.now().strftime('%d/%m/%Y %H:%M')} · {len(data)} registros"
        ws.cell(row=2, column=1, value=sub).font = subtitulo_font

        ws.append([])  # fila 3 en blanco
        fila_header = 4
        ws.append(headers)  # fila 4: encabezados

        for cell in ws[fila_header]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
        ws.row_dimensions[fila_header].height = 22

        for i, fila in enumerate(data, start=1):
            ws.append([_formatear_valor(fila.get(h), h) for h in headers_raw])
            fila_actual = fila_header + i
            for cell in ws[fila_actual]:
                cell.border = borde_fino
                if i % 2 == 0:
                    cell.fill = zebra_fill

        ws.freeze_panes = ws.cell(row=fila_header + 1, column=1)
        ultima_col_letra = ws.cell(row=fila_header, column=n_cols).column_letter
        ws.auto_filter.ref = f"A{fila_header}:{ultima_col_letra}{fila_header}"

        # Autoajuste de ancho usando el valor YA formateado
        for idx, header in enumerate(headers, start=1):
            col_letter = ws.cell(row=fila_header, column=idx).column_letter
            max_length = len(header)
            for fila in data:
                valor = _formatear_valor(fila.get(headers_raw[idx - 1]), headers_raw[idx - 1])
                max_length = max(max_length, len(str(valor)))
            ws.column_dimensions[col_letter].width = min(max(max_length + 4, 12), 40)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    response = HttpResponse(output.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    return response


def _pdf_header_footer_kpi(titulo_reporte, fecha_inicio=None, fecha_fin=None):
    """Header/footer con logo y marca, mismo estilo que ordenes de mantenimiento."""
    def _draw(canvas, doc):
        canvas.saveState()
        margin = 20 * mm

        logo_w = 0
        if _logo_drawing is not None:
            logo_w = 14 * mm
            try:
                escala = logo_w / _logo_drawing.width
                canvas.saveState()
                canvas.translate(margin, PAGE_H_KPI - margin - logo_w + 6 * mm)
                canvas.scale(escala, escala)
                renderPDF.draw(_logo_drawing, canvas, 0, 0)
                canvas.restoreState()
            except Exception:
                logo_w = 0

        text_x = margin + (logo_w + 4 * mm if logo_w else 0)
        canvas.setFillColor(BRAND_DARK)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawString(text_x, PAGE_H_KPI - margin, "OperaCore")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(BRAND_MUTED)
        canvas.drawString(text_x, PAGE_H_KPI - margin - 10, "Reporte de indicadores")

        canvas.setFont("Helvetica-Bold", 12)
        canvas.setFillColor(BRAND_DARK)
        canvas.drawRightString(PAGE_W_KPI - margin, PAGE_H_KPI - margin, titulo_reporte)

        if fecha_inicio or fecha_fin:
            rango = f"Periodo: {fecha_inicio or 'inicio'} — {fecha_fin or 'hoy'}"
            canvas.setFont("Helvetica", 8)
            canvas.setFillColor(BRAND_MUTED)
            canvas.drawRightString(PAGE_W_KPI - margin, PAGE_H_KPI - margin - 10, rango)

        canvas.setStrokeColor(BRAND_ACCENT)
        canvas.setLineWidth(1.3)
        line_y = PAGE_H_KPI - margin - 20 * mm
        canvas.line(margin, line_y, PAGE_W_KPI - margin, line_y)

        canvas.setStrokeColor(BRAND_LINE)
        canvas.setLineWidth(0.6)
        canvas.line(margin, margin - 4 * mm, PAGE_W_KPI - margin, margin - 4 * mm)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(BRAND_MUTED)
        generado = timezone.now().strftime("%d/%m/%Y %H:%M")
        canvas.drawString(margin, margin - 10 * mm, "OperaCore · Documento generado automáticamente el " + generado)
        canvas.drawRightString(PAGE_W_KPI - margin, margin - 10 * mm, "Página " + str(doc.page))

        canvas.restoreState()
    return _draw


def _pdf_titulo_seccion(titulo):
    """Barra de acento + título, mismo lenguaje visual que las secciones
    de la orden de mantenimiento."""
    style = ParagraphStyle("kpiSeccion", fontName="Helvetica-Bold", fontSize=12,
                            textColor=BRAND_DARK, leading=14)
    barra = Table([[""]], colWidths=[3], rowHeights=[16])
    barra.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), BRAND_ACCENT)]))
    cuerpo = Table([[barra, Paragraph(titulo, style)]], colWidths=[3, 300], hAlign="LEFT")
    cuerpo.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("LEFTPADDING", (1, 0), (1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    return cuerpo


def _pdf_calcular_anchos(headers, headers_raw, data):
    """Ancho por columna basado en el contenido real (no fuerza a llenar
    toda la pagina si la tabla es angosta, ej. Estado de la flota)."""
    anchos = []
    for h, h_raw in zip(headers, headers_raw):
        muestra = [_formatear_valor(fila.get(h_raw), h_raw) for fila in data[:80]]
        max_chars = max([len(h)] + [len(v) for v in muestra])
        ancho = max_chars * 1.9 * mm + 6 * mm
        ancho = min(max(ancho, 24 * mm), 65 * mm)
        anchos.append(ancho)

    ancho_disponible = PAGE_W_KPI - 40 * mm
    total = sum(anchos)
    if total > ancho_disponible:
        factor = ancho_disponible / total
        anchos = [a * factor for a in anchos]
    return anchos


def _pdf_tabla_kpi(data):
    """Tabla con celdas tipo Paragraph (para que el texto largo haga wrap
    en vez de desbordarse), encabezado repetido si salta de pagina, y
    ancho de columna ajustado al contenido real."""
    header_style = ParagraphStyle("kpiHeaderCell", fontName="Helvetica-Bold", fontSize=8,
                                   textColor=colors.white, leading=10)
    cell_style = ParagraphStyle("kpiBodyCell", fontName="Helvetica", fontSize=8,
                                 textColor=colors.HexColor("#1f2937"), leading=10)

    headers_raw = list(data[0].keys())
    headers = [_humanizar_columna(h) for h in headers_raw]

    table_data = [[Paragraph(h, header_style) for h in headers]]
    for fila in data:
        table_data.append([
            Paragraph(_formatear_valor(fila.get(h), h), cell_style) for h in headers_raw
        ])

    col_widths = _pdf_calcular_anchos(headers, headers_raw, data)

    tabla = Table(table_data, colWidths=col_widths, repeatRows=1, hAlign="LEFT")
    zebra = colors.HexColor("#f8fafc")
    estilos = [
        ("BACKGROUND", (0, 0), (-1, 0), BRAND_ACCENT),
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("TOPPADDING", (0, 0), (-1, -1), 5),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
        ("LEFTPADDING", (0, 0), (-1, -1), 6),
        ("GRID", (0, 0), (-1, -1), 0.4, BRAND_LINE),
    ]
    for i in range(2, len(table_data), 2):
        estilos.append(("BACKGROUND", (0, i), (-1, i), zebra))
    tabla.setStyle(TableStyle(estilos))
    return tabla


def _generar_pdf(vistas_data, nombre_archivo="reporte_kpi", fecha_inicio=None, fecha_fin=None):
    """Genera un PDF con una seccion por vista, con header/footer de marca."""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer, pagesize=landscape(letter),
        topMargin=46 * mm, bottomMargin=26 * mm,
        leftMargin=20 * mm, rightMargin=20 * mm,
        title="Reporte de indicadores OperaCore",
    )

    vistas_con_datos = [(v, d) for v, d in vistas_data if d]

    story = []
    for idx, (vista, data) in enumerate(vistas_con_datos):
        story.append(_pdf_titulo_seccion(TITULOS_KPI.get(vista, vista)))
        story.append(Spacer(1, 8))
        story.append(_pdf_tabla_kpi(data))
        if idx < len(vistas_con_datos) - 1:
            story.append(PageBreak())

    if not story:
        story = [Paragraph(
            "No hay datos para el periodo y las vistas seleccionadas.",
            getSampleStyleSheet()["Normal"],
        )]

    if len(vistas_con_datos) == 1:
        titulo_reporte = TITULOS_KPI.get(vistas_con_datos[0][0], "Reporte de indicadores")
    else:
        titulo_reporte = "Reporte de indicadores"

    draw = _pdf_header_footer_kpi(titulo_reporte, fecha_inicio, fecha_fin)
    doc.build(story, onFirstPage=draw, onLaterPages=draw)

    buffer.seek(0)
    response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="{nombre_archivo}_{timezone.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    return response


class ReporteKPIExportAPIView(APIView):
    """Exporta una o mas vistas de KPI a CSV, Excel o PDF.
    GET /indicadores/v1/reporte/export/<formato>/?vistas=slug1,slug2&fecha_inicio=...&fecha_fin=...

    - formato: csv, xlsx o pdf
    - vistas: lista de slugs de vistas (ej: vistas=estado-flota,top-fallas)
    - fecha_inicio, fecha_fin: opcionales, filtro de periodo (YYYY-MM-DD)
    """

    def get(self, request, formato):
        vistas_slugs = request.query_params.get("vistas", "").split(",")
        vistas_slugs = [v.strip() for v in vistas_slugs if v.strip()]

        if not vistas_slugs:
            return Response(
                {"detail": "Debe especificar al menos una vista con el parametro 'vistas'."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Validar vistas
        vistas_invalidas = [v for v in vistas_slugs if v not in VISTAS_KPI]
        if vistas_invalidas:
            return Response(
                {
                    "detail": f"Vistas invalidas: {', '.join(vistas_invalidas)}. Disponibles: {', '.join(VISTAS_KPI.keys())}"
                },
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Parsear fechas
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")

        try:
            if fecha_inicio:
                fecha_inicio = timezone.datetime.fromisoformat(fecha_inicio).date()
            if fecha_fin:
                fecha_fin = timezone.datetime.fromisoformat(fecha_fin).date()
        except ValueError:
            return Response(
                {"detail": "Formato de fecha invalido. Use YYYY-MM-DD."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Obtener datos de cada vista
        vistas_data = []
        for slug in vistas_slugs:
            tabla = VISTAS_KPI[slug]
            with connection.cursor() as cur:
                cur.execute(f"SELECT * FROM {tabla}")
                data = _filas_a_dicts(cur)

                # Filtrar por periodo
                if fecha_inicio or fecha_fin:
                    data = [fila for fila in data if _fila_en_periodo(fila, fecha_inicio, fecha_fin)]

                vistas_data.append((slug, data))

        # Generar archivo
        if formato == "csv":
            return _generar_csv(vistas_data)
        elif formato == "xlsx":
            return _generar_xlsx(vistas_data)
        elif formato == "pdf":
            return _generar_pdf(vistas_data, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin)
        else:
            return Response(
                {"detail": "Formato no soportado. Use csv, xlsx o pdf."},
                status=status.HTTP_400_BAD_REQUEST,
            )


class CerrarPeriodoIndicadorAPIView(APIView):
    """Cierra el periodo vigente de una maquina en INDICADOR y abre el
    siguiente, via sp_cerrar_periodo_indicador. Boton manual del panel de
    indicadores: admin elige maquina + fecha.
    Body: {"maquina": "MAQ001", "fecha_fin": "2026-02-28"}"""

    def post(self, request):
        maquina = request.data.get("maquina")
        fecha_fin = request.data.get("fecha_fin")
        if not maquina or not fecha_fin:
            return Response(
                {"detail": "Faltan 'maquina' y/o 'fecha_fin'."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            with connection.cursor() as cur:
                cur.callproc("sp_cerrar_periodo_indicador", [maquina, fecha_fin])
        except OperationalError as e:
            # Los SIGNAL SQLSTATE '45000' del SP (maquina no existe, no hay
            # periodo abierto, fecha invalida) llegan aqui.
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(
            {"detail": "Periodo cerrado y nuevo periodo abierto.", "maquina": maquina, "fecha_fin": fecha_fin},
            status=status.HTTP_200_OK,
        )


class ReporteDisponibilidadPlantaAPIView(APIView):
    """Reporte de disponibilidad/MTBF/MTTR + fallas/ordenes por linea para
    un rango de fechas arbitrario, via sp_reporte_disponibilidad_planta.
    GET /indicadores/v1/reporte-disponibilidad/?fecha_inicio=...&fecha_fin=..."""

    def get(self, request):
        fecha_inicio = request.query_params.get("fecha_inicio")
        fecha_fin = request.query_params.get("fecha_fin")
        if not fecha_inicio or not fecha_fin:
            return Response(
                {"detail": "Faltan 'fecha_inicio' y/o 'fecha_fin' como query params."},
                status=status.HTTP_400_BAD_REQUEST,
            )
        try:
            with connection.cursor() as cur:
                cur.callproc("sp_reporte_disponibilidad_planta", [fecha_inicio, fecha_fin])
                data = _filas_a_dicts(cur)
        except OperationalError as e:
            return Response({"detail": str(e)}, status=status.HTTP_400_BAD_REQUEST)

        return Response(data, status=status.HTTP_200_OK)


class PingAPIView(APIView):
    """Endpoint de prueba: confirma que el modulo Indicadores responde."""

    def get(self, request):
        return Response({"modulo": "indicadores", "status": "ok"}, status=status.HTTP_200_OK)


class ResumenIndicadoresAPIView(APIView):
    """Resumen para el Dashboard (RF-26): MTBF, MTTR y disponibilidad
    promedio de la flota (tomando el ultimo registro de INDICADOR por
    maquina), ordenes pendientes y alertas de inventario (refacciones
    en o debajo de su stock minimo)."""

    def get(self, request):
        # MySQL no soporta distinct("campo") como Postgres, asi que se
        # resuelve a mano: nos quedamos con el INDICADOR mas reciente
        # (numeroRegistro mayor) de cada maquina.
        ultimo_por_maquina = {}
        for ind in Indicador.objects.select_related("maquina").order_by("maquina", "-numeroRegistro"):
            if ind.maquina_id not in ultimo_por_maquina:
                ultimo_por_maquina[ind.maquina_id] = ind

        indicadores = list(ultimo_por_maquina.values())

        def promedio(campo):
            valores = [getattr(i, campo) for i in indicadores if getattr(i, campo) is not None]
            return round(sum(valores) / len(valores), 1) if valores else None

        # Detalle por maquina para la tablita del dashboard.
        por_maquina = []
        for ind in sorted(indicadores, key=lambda i: i.maquina_id or ""):
            nombre_maquina = ind.maquina.nombre if ind.maquina else ind.maquina_id
            por_maquina.append({
                "codigo": ind.maquina_id,
                "nombre": nombre_maquina,
                "mtbf": ind.mtbf,
                "mttr": ind.mttr,
                "disponibilidad": ind.porcentajeDispo,
            })

        data = {
            "mtbf_promedio": promedio("mtbf"),
            "mttr_promedio": promedio("mttr"),
            "disponibilidad_promedio": promedio("porcentajeDispo"),
            "maquinas_con_indicador": len(indicadores),
            "por_maquina": por_maquina,
            "ordenes_pendientes": OrdenMantenimiento.objects.filter(
                estado_orden_id="PENDI"
            ).count(),
            "alertas_inventario": Refaccion.objects.filter(
                stock__lte=F("stockminimo")
            ).count(),
        }
        return Response(data, status=status.HTTP_200_OK)


# A partir de aqui sigue el patron de tu maestro cuando agregues modelos reales:
#
# class ListIndicadoresAPIView(generics.ListAPIView):
#     queryset = models.MiModelo.objects.all()
#     serializer_class = serializers.ListMaquinaSerializer
#
# class DetailIndicadoresAPIView(generics.RetrieveAPIView):
#     queryset = models.MiModelo.objects.all()
#     serializer_class = serializers.DetailMaquinaSerializer
#
# class CreateIndicadoresAPIView(generics.CreateAPIView):
#     serializer_class = serializers.CreateMaquinaSerializer
#
# class UpdateIndicadoresAPIView(generics.UpdateAPIView):
#     queryset = models.MiModelo.objects.all()
#     serializer_class = serializers.CreateMaquinaSerializer
#
# class DeleteIndicadoresAPIView(generics.DestroyAPIView):
#     queryset = models.MiModelo.objects.all()