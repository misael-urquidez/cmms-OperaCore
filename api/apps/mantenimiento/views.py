import csv
import os
import io

from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Font
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak
from rest_framework import generics, status
from rest_framework.exceptions import NotFound, PermissionDenied, ValidationError
from rest_framework.response import Response
from rest_framework.views import APIView
from apps.usuarios.models import Trabajador
from apps.maquinaria.services import cambiar_estado_maquina
from apps.maquinaria.models import Maquina
from django.db import connection, transaction
from django.db.utils import OperationalError
from . import models
from . import serializers
from apps.fallas.views import _get_reporte_data
from apps.inventario.stock import liberar_herramienta_de_orden


class PingAPIView(APIView):
    """Endpoint de prueba: confirma que el modulo Mantenimiento responde."""

    def get(self, request):
        return Response({"modulo": "mantenimiento", "status": "ok"}, status=status.HTTP_200_OK)



# ------------ ESTADO_ORDEN --------------------------------------------
class EstadoOrdenListAPIView(generics.ListAPIView):
    queryset = models.EstadoOrden.objects.all().order_by("nombre")
    serializer_class = serializers.ListEstadoOrdenSerializer


class EstadoOrdenCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateEstadoOrdenSerializer


class EstadoOrdenDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.EstadoOrden.objects.all()
    lookup_field = "codigo"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.CreateEstadoOrdenSerializer
        return serializers.DetailEstadoOrdenSerializer


# ------------ TIPO_MANTENIMIENTO ---------------------------------------
class TipoMantenimientoListAPIView(generics.ListAPIView):
    queryset = models.TipoMantenimiento.objects.all().order_by("nombre")
    serializer_class = serializers.ListTipoMantenimientoSerializer


class TipoMantenimientoCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTipoMantenimientoSerializer


class TipoMantenimientoDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.TipoMantenimiento.objects.all()
    lookup_field = "codigo"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.CreateTipoMantenimientoSerializer
        return serializers.DetailTipoMantenimientoSerializer


# ------------ TAREAS -----------------------------------------------------
class TareasListAPIView(generics.ListAPIView):
    queryset = models.Tareas.objects.all().order_by("numeroregistro")
    serializer_class = serializers.ListTareasSerializer


class TareasCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTareasSerializer


class TareasDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.Tareas.objects.all()
    lookup_field = "numeroregistro"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.CreateTareasSerializer
        return serializers.DetailTareasSerializer


# ------------ TIPO_MOVIMIENTO -------------------------------------------
class TipoMovimientoListAPIView(generics.ListAPIView):
    queryset = models.TipoMovimiento.objects.all().order_by("codigo")
    serializer_class = serializers.ListTipoMovimientoSerializer


class TipoMovimientoCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTipoMovimientoSerializer


class TipoMovimientoDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.TipoMovimiento.objects.all()
    lookup_field = "codigo"

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.CreateTipoMovimientoSerializer
        return serializers.DetailTipoMovimientoSerializer


# ------------ MOVIMIENTO ---------------------------------------------------
class MovimientoListAPIView(generics.ListAPIView):
    queryset = models.Movimiento.objects.select_related(
        "orden_mantenimiento", "refaccion", "pieza",
    ).order_by("-fecha", "-hora")
    serializer_class = serializers.ListMovimientoSerializer


class MovimientoCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateMovimientoSerializer

    def create(self, request, *args, **kwargs):
        data = request.data.copy()
        tipo = data.get("tipoMovimiento")

        with transaction.atomic():
            if tipo == "INSTA":
                movimiento = self._registrar_insta(data)
            elif tipo == "REHA":
                movimiento = self._registrar_reha(data)
            elif tipo == "DESMO":
                movimiento = self._registrar_desmo(data)
            else:
                # Cualquier otro tipo: solo se deja el registro en MOVIMIENTO.
                data.pop("pieza_data", None)
                data.pop("refaccion_data", None)
                serializer = self.get_serializer(data=data)
                serializer.is_valid(raise_exception=True)
                movimiento = serializer.save()

        return Response(
            serializers.ListMovimientoSerializer(movimiento).data,
            status=status.HTTP_201_CREATED,
        )

    # ------------------------------------------------------------------
    def _crear_pieza_nueva(self, data):
        """Crea la pieza a partir de pieza_data (INSTA). Rechaza con mensaje
        claro si el numeroSerie ya existe para no duplicar una unidad fisica."""
        from apps.inventario.serializers import CreatePiezaSerializer

        pieza_data = data.get("pieza_data") or None
        if not pieza_data:
            return None, data.get("pieza") or None

        pieza_campos = dict(pieza_data)
        for k in ("pieza", "pieza_data"):
            pieza_campos.pop(k, None)

        serie = (pieza_campos.get("numeroserie") or "").strip()
        if serie and models.Pieza.objects.filter(numeroserie=serie).exists():
            raise ValidationError(
                f"La serie {serie} ya existe. Si la pieza quedó desmontada, "
                "úsala en DESMO/REHA; no vuelvas a registrarla."
            )

        if not pieza_campos.get("nombre") and data.get("refaccion"):
            ref = models.Refaccion.objects.filter(pk=data["refaccion"]).first()
            if ref:
                pieza_campos["nombre"] = ref.nombre
                pieza_campos.setdefault("costoinicial", ref.costo)
        if not pieza_campos.get("fechainstalacion") and data.get("fecha"):
            pieza_campos["fechainstalacion"] = data["fecha"]
        if not pieza_campos.get("maquina") and data.get("orden_mantenimiento"):
            orden = models.OrdenMantenimiento.objects.filter(
                pk=data["orden_mantenimiento"]
            ).first()
            if orden and orden.maquina_id:
                pieza_campos["maquina"] = orden.maquina_id

        pieza_ser = CreatePiezaSerializer(data=pieza_campos)
        pieza_ser.is_valid(raise_exception=True)
        pieza = pieza_ser.save()
        return pieza, pieza.numeroserie

    def _call_sp_salida_refaccion(self, data, pieza_serie):
        """Ejecuta sp_registrar_salida_refaccion (DISPO -> INMAQ + registro
        MOVIMIENTO INSTA). Los SIGNAL del SP llegan como OperationalError y se
        convierten en un 400; el transaction.atomic revierte todo."""
        try:
            with connection.cursor() as cur:
                cur.callproc(
                    "sp_registrar_salida_refaccion",
                    [
                        data.get("refaccion"),
                        data.get("orden_mantenimiento"),
                        data.get("descripcion"),
                        data.get("fecha"),
                        data.get("hora"),
                        pieza_serie,
                    ],
                )
                col_names = [c[0] for c in cur.description]
                row = cur.fetchone()
                while cur.nextset():
                    pass
                if not row:
                    raise ValidationError(
                        "El procedimiento no devolvió un resultado."
                    )
                return dict(zip(col_names, row))
        except OperationalError as e:
            raise ValidationError(e.args[1] if len(e.args) > 1 else str(e))

    def _registrar_insta(self, data):
        """INSTA: la refaccion sale del almacen (SP3 DISPO -> INMAQ) y la
        pieza recien creada (siempre pieza nueva) queda instalada en la
        maquina de la orden."""
        from apps.inventario.models import EdoPieza

        if not data.get("pieza_data"):
            raise ValidationError(
                "INSTA requiere registrar la pieza nueva que se instala."
            )
        pieza, pieza_serie = self._crear_pieza_nueva(data)

        resultado = self._call_sp_salida_refaccion(data, pieza_serie)
        movimiento = models.Movimiento.objects.get(pk=resultado.get("numero_movimiento"))

        # La pieza recién creada queda OPERA en la maquina de la orden. Si el
        # UPDATE fallara, el atomic revierte el SP y la pieza.
        if pieza is not None:
            opera = EdoPieza.objects.filter(codigo="OPERA").first()
            if opera:
                pieza.edo_pieza_id = "OPERA"
            if data.get("orden_mantenimiento"):
                orden = models.OrdenMantenimiento.objects.filter(
                    pk=data["orden_mantenimiento"]
                ).first()
                if orden and orden.maquina_id:
                    pieza.maquina_id = orden.maquina_id
            pieza.save(update_fields=["maquina", "edo_pieza"])

        return movimiento

    def _registrar_reha(self, data):
        """REHA: una pieza desmontada y en ENREH (sin maquina) se rehabilita:
        queda OPERA y, si la orden seleccionada trae maquina, se reasigna a esa
        maquina (igual que INSTA). No se crea ni modifica ninguna refaccion:
        la pieza fisica conserva su registro."""
        import datetime

        pieza_serie = data.get("pieza") or None
        if not pieza_serie:
            raise ValidationError("Indica la pieza que se está rehabilitando.")

        pieza = models.Pieza.objects.filter(numeroserie=pieza_serie).first()
        if not pieza:
            raise ValidationError(f"La pieza {pieza_serie} no existe.")
        if pieza.maquina_id is not None:
            raise ValidationError(
                f"La pieza {pieza_serie} sigue instalada en una máquina; "
                "usa DESMO antes de rehabilitarla."
            )
        if pieza.edo_pieza_id != "ENREH":
            raise ValidationError(
                "La pieza debe estar en estado 'En Rehabilitación' (ENREH) "
                "para rehabilitarse."
            )

        pieza.edo_pieza_id = "OPERA"
        if data.get("orden_mantenimiento"):
            orden = models.OrdenMantenimiento.objects.filter(
                pk=data["orden_mantenimiento"]
            ).first()
            if orden and orden.maquina_id:
                pieza.maquina_id = orden.maquina_id
        pieza.save(update_fields=["maquina", "edo_pieza"])

        return models.Movimiento.objects.create(
            descripcion=data.get("descripcion")
            or f"Rehabilitación de pieza {pieza.numeroserie}",
            fecha=data.get("fecha") or datetime.date.today(),
            hora=data.get("hora") or datetime.datetime.now().time(),
            tipomovimiento="REHA",
            orden_mantenimiento_id=data.get("orden_mantenimiento") or None,
            pieza=pieza,
        )

    def _registrar_desmo(self, data):
        """DESMO: la pieza sale de su maquina (maquina = NULL) y pasa a
        'En Rehabilitación' (ENREH), el estado previo requerido por REHA."""
        import datetime

        pieza_serie = data.get("pieza") or None
        if not pieza_serie:
            raise ValidationError("Indica la pieza que se está desmontando.")

        pieza = models.Pieza.objects.filter(numeroserie=pieza_serie).first()
        if not pieza:
            raise ValidationError(f"La pieza {pieza_serie} no existe.")
        if pieza.maquina_id is None:
            raise ValidationError(
                f"La pieza {pieza_serie} no está instalada en ninguna máquina."
            )

        maquina_origen = pieza.maquina
        pieza.maquina = None
        if pieza.edo_pieza_id in ("OPERA", "DEGRA", "FALLI"):
            pieza.edo_pieza_id = "ENREH"
        pieza.save(update_fields=["maquina", "edo_pieza"])

        return models.Movimiento.objects.create(
            descripcion=data.get("descripcion")
            or (
                f"Pieza retirada de la máquina {maquina_origen.nombre}"
                if maquina_origen
                else "Pieza desmontada"
            ),
            fecha=data.get("fecha") or datetime.date.today(),
            hora=data.get("hora") or datetime.datetime.now().time(),
            tipomovimiento="DESMO",
            orden_mantenimiento_id=data.get("orden_mantenimiento") or None,
            pieza=pieza,
        )


# ------------ TAREA_ORDEN (llave compuesta) ------------------------------
class TareaOrdenListAPIView(generics.ListAPIView):
    queryset = models.TareaOrden.objects.all()
    serializer_class = serializers.ListTareaOrdenSerializer


class TareaOrdenCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTareaOrdenSerializer


class TareaOrdenDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.TareaOrden.objects.all()
    serializer_class = serializers.DetailTareaOrdenSerializer

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.UpdateTareaOrdenSerializer
        return serializers.DetailTareaOrdenSerializer

    def get_object(self):
        obj = generics.get_object_or_404(
            self.get_queryset(),
            tarea=self.kwargs["tarea"],
            orden_mantenimiento=self.kwargs["orden_mantenimiento"],
        )
        self.check_object_permissions(self.request, obj)
        return obj

    def perform_update(self, serializer):
        orden = serializer.instance.orden_mantenimiento
        if orden.estado_orden_id != "ENPRO":
            raise PermissionDenied("Solo se pueden marcar tareas con la orden en progreso (ENPRO).")
        # TAREA_ORDEN tiene llave compuesta (tarea, orden_mantenimiento) que
        # Django no sabe modelar: save() sobre la instancia cae en un INSERT
        # (pk duplicada). Se actualiza por queryset, igual que en el resto
        # del modulo, y el porcentaje se recalcula despues.
        models.TareaOrden.objects.filter(
            tarea=serializer.instance.tarea_id,
            orden_mantenimiento=orden,
        ).update(**serializer.validated_data)
        serializers._recalcular_porcentaje(orden)


# ------------ HERRA_ORDEN (llave compuesta) -------------------------------
class HerraOrdenListAPIView(generics.ListAPIView):
    queryset = models.HerraOrden.objects.all()
    serializer_class = serializers.ListHerraOrdenSerializer


class HerraOrdenCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateHerraOrdenSerializer


class HerraOrdenDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.HerraOrden.objects.all()
    serializer_class = serializers.DetailHerraOrdenSerializer

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.UpdateHerraOrdenSerializer
        return serializers.DetailHerraOrdenSerializer

    def perform_destroy(self, instance):
        liberar_herramienta_de_orden(instance.herramienta_id)
        # PK compuesta emulada: delete() del objeto usaria solo "herramienta"
        # en el WHERE y borraria las filas de TODAS las ordenes. Se borra con
        # la llave completa.
        models.HerraOrden.objects.filter(
            herramienta=instance.herramienta_id,
            orden_mantenimiento=instance.orden_mantenimiento_id,
        ).delete()

    def get_object(self):
        obj = generics.get_object_or_404(
            self.get_queryset(),
            herramienta=self.kwargs["herramienta"],
            orden_mantenimiento=self.kwargs["orden_mantenimiento"],
        )
        self.check_object_permissions(self.request, obj)
        return obj


# ------------ TRABA_ORDE_PERSONAL (llave compuesta) -----------------------
class TrabaOrdePersonalListAPIView(generics.ListAPIView):
    queryset = models.TrabaOrdePersonal.objects.all()
    serializer_class = serializers.ListTrabaOrdePersonalSerializer


class TrabaOrdePersonalCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateTrabaOrdePersonalSerializer


class TrabaOrdePersonalDetailAPIView(generics.RetrieveUpdateDestroyAPIView):
    queryset = models.TrabaOrdePersonal.objects.all()
    serializer_class = serializers.DetailTrabaOrdePersonalSerializer

    def get_serializer_class(self):
        if self.request.method in ("PUT", "PATCH"):
            return serializers.UpdateTrabaOrdePersonalSerializer
        return serializers.DetailTrabaOrdePersonalSerializer

    def get_object(self):
        obj = generics.get_object_or_404(
            self.get_queryset(),
            trabajador=self.kwargs["trabajador"],
            orden_mantenimiento=self.kwargs["orden_mantenimiento"],
        )
        self.check_object_permissions(self.request, obj)
        return obj

# A partir de aqui sigue el mismo patron cuando quieras exponer
# OrdenMantenimiento / Movimiento.


# ------------ ORDEN_MANTENIMIENTO ---------------------------------------
# ------------ REPORTE_FALLA disponibles (para "adjuntar reporte") --------
class ReporteFallaDisponibleListAPIView(generics.ListAPIView):
    """Reportes de falla elegibles para adjuntarse a una nueva orden correctiva:
    de la maquina indicada, en un estado 'vivo' (ABIER/ENATE/ENESP) y que
    todavia no tengan ninguna orden de mantenimiento vinculada."""
    serializer_class = serializers.ReporteFallaDisponibleSerializer

    def get_queryset(self):
        from apps.fallas.models import ReporteFalla
        maquina = self.request.query_params.get("maquina")
        if not maquina:
            return ReporteFalla.objects.none()
        qs = ReporteFalla.objects.select_related("tipo_severidad", "estado_reporte").filter(
            maquina_id=maquina,
            estado_reporte_id__in=["ABIER", "ENATE", "ENESP"],
        ).exclude(mantenimiento_ordenes__isnull=False)
        return qs.order_by("-fechaCreacion", "-horaCreacion")


class OrdenMantenimientoListAPIView(generics.ListAPIView):
    serializer_class = serializers.ListOrdenMantenimientoSerializer
    def get_queryset(self):
        qs = models.OrdenMantenimiento.objects.select_related("maquina", "trabajador", "tipo_mantenimiento", "estado_orden", "reporte_falla").order_by("-fechacreacion", "-horacreacion")
        if self.request.query_params.get("trabajador"):
            qs = qs.filter(trabajador_id=self.request.query_params["trabajador"])
        if self.request.query_params.get("estado"):
            qs = qs.filter(estado_orden_id=self.request.query_params["estado"])
        if self.request.query_params.get("tipo_mantenimiento"):
            qs = qs.filter(tipo_mantenimiento_id=self.request.query_params["tipo_mantenimiento"])
        return qs

class OrdenMantenimientoDetailAPIView(generics.RetrieveAPIView):
    queryset = models.OrdenMantenimiento.objects.select_related("maquina", "trabajador", "tipo_mantenimiento", "estado_orden")
    serializer_class = serializers.DetailOrdenMantenimientoSerializer
    lookup_field = "folio"

class OrdenMantenimientoCreateAPIView(generics.CreateAPIView):
    serializer_class = serializers.CreateOrdenMantenimientoSerializer
    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data); serializer.is_valid(raise_exception=True)
        return Response(serializers.DetailOrdenMantenimientoSerializer(serializer.save()).data, status=status.HTTP_201_CREATED)

class OrdenMantenimientoAsignarAPIView(APIView):
    def patch(self, request, folio):
        try: orden = models.OrdenMantenimiento.objects.get(pk=folio)
        except models.OrdenMantenimiento.DoesNotExist as exc: raise NotFound("Orden no encontrada.") from exc
        serializer = serializers.AsignarTrabajadorOrdenSerializer(data=request.data); serializer.is_valid(raise_exception=True)
        try: orden.trabajador = Trabajador.objects.get(pk=serializer.validated_data["trabajador"])
        except Trabajador.DoesNotExist: return Response({"trabajador": "No existe ese trabajador."}, status=400)
        # Una orden correctiva nace como SOLIC (solicitada por falla, no
        # programada) y se queda asi aunque se le asigne responsable: no debe
        # volverse "Programada". Solo las ordenes planificadas (preventivo,
        # predictivo, etc.) pasan SOLIC -> PROGR al asignar.
        if orden.estado_orden_id == "SOLIC" and orden.tipo_mantenimiento_id != "CORRE":
            orden.estado_orden_id = "PROGR"
        orden.save(update_fields=["trabajador", "estado_orden"])
        return Response(serializers.DetailOrdenMantenimientoSerializer(orden).data)

class OrdenMantenimientoIniciarAPIView(APIView):
    @transaction.atomic
    def patch(self, request, folio):
        try: orden = models.OrdenMantenimiento.objects.get(pk=folio)
        except models.OrdenMantenimiento.DoesNotExist as exc: raise NotFound("Orden no encontrada.") from exc
        orden.estado_orden_id = "ENPRO"; orden.save(update_fields=["estado_orden"])
        # El diagrama solo contempla OPERA -> FALLO -> MANTE: la maquina nada
        # mas se mueve a MANTE cuando de verdad viene de una falla reportada.
        # Preventivo/predictivo/emergencia (sin reporte_falla) no representan
        # un paro no planificado, asi que la maquina se queda como estaba y
        # solo se lleva el seguimiento en la propia orden.
        maquina = Maquina.objects.get(pk=orden.maquina_id)
        if maquina.estado_maquina_id == "FALLO":
            cambiar_estado_maquina(orden.maquina_id, "MANTE", "orden_mantenimiento", orden.folio)
        return Response(serializers.DetailOrdenMantenimientoSerializer(orden).data)

class OrdenMantenimientoCerrarAPIView(APIView):
    @transaction.atomic
    def patch(self, request, folio):
        try: orden = models.OrdenMantenimiento.objects.get(pk=folio)
        except models.OrdenMantenimiento.DoesNotExist as exc: raise NotFound("Orden no encontrada.") from exc
        if orden.fechacierre is not None: return Response({"detail": "Esta orden ya está cerrada."}, status=400)
        serializer = serializers.CerrarOrdenSerializer(data=request.data); serializer.is_valid(raise_exception=True)
        datos, ahora = serializer.validated_data, timezone.localtime()
        for campo, valor in ((("diagnostico", datos.get("diagnostico")), ("notas", datos.get("notas")), ("horasintervenidas", datos.get("horasIntervenidas")))):
            if valor is not None: setattr(orden, campo, valor)

        # IMPORTANTE: cerrar el reporte de falla ANTES de guardar fechaCierre
        # en la orden, para que tg_actualizar_mttr_orden (dispara con el
        # UPDATE de abajo) lea el tiempoParo ya actualizado. Ambos saves
        # van en la misma transacción gracias al @transaction.atomic.
        if orden.reporte_falla_id:
            from datetime import datetime
            reporte = orden.reporte_falla
            creado = datetime.combine(reporte.fechaCreacion, reporte.horaCreacion)
            reporte.tiempoParo = int((ahora.replace(tzinfo=None) - creado).total_seconds() / 3600)
            reporte.estado_reporte_id = "RESUE"
            reporte.save(update_fields=["tiempoParo", "estado_reporte"])

        orden.fechacierre, orden.horacierre, orden.estado_orden_id = ahora.date(), ahora.time(), "CERRA"
        orden.save(update_fields=["diagnostico", "notas", "horasintervenidas", "fechacierre", "horacierre", "estado_orden"])

        # Las herramientas asignadas a la orden regresan a DISPO (disponibles).
        for ho in models.HerraOrden.objects.filter(orden_mantenimiento=orden):
            liberar_herramienta_de_orden(ho.herramienta_id)

        # Movimientos de inventario capturados en el drawer de cierre. Por
        # cada renglon: si trae "pieza" (la pieza fisica que salio de la
        # maquina) se emite un DESMO para esa pieza, y SIEMPRE se emite un
        # INSTA para la refaccion que entro. Antes se guardaba un solo INSTA
        # mezclando refaccion nueva + pieza vieja, y la pieza retirada nunca
        # quedaba registrada como "salida" en el rastro de Inventario.
        for item in datos.get("movimientos") or []:
            refaccion = item["refaccion"]
            pieza = item.get("pieza")
            if pieza is not None:
                models.Movimiento.objects.create(
                    descripcion=f"Pieza retirada al cerrar orden {orden.folio}",
                    fecha=ahora.date(),
                    hora=ahora.time(),
                    tipomovimiento="DESMO",
                    orden_mantenimiento=orden,
                    pieza=pieza,
                )
            models.Movimiento.objects.create(
                descripcion=f"Refacción instalada al cerrar orden {orden.folio}",
                fecha=ahora.date(),
                hora=ahora.time(),
                tipomovimiento="INSTA",
                orden_mantenimiento=orden,
                refaccion=refaccion,
            )

        # Simetrico al guard de "iniciar": si la maquina nunca entro a MANTE
        # (preventivo/predictivo/emergencia que no tocaron su estado) no hay
        # nada que regresar a ESPER; se deja como esta.
        maquina = Maquina.objects.get(pk=orden.maquina_id)
        if maquina.estado_maquina_id == "MANTE":
            cambiar_estado_maquina(orden.maquina_id, "ESPER", "orden_mantenimiento", orden.folio)
        return Response(serializers.DetailOrdenMantenimientoSerializer(orden).data)


class OrdenMantenimientoUpdateAPIView(APIView):
    def patch(self, request, folio):
        try:
            orden = models.OrdenMantenimiento.objects.get(pk=folio)
        except models.OrdenMantenimiento.DoesNotExist as exc:
            raise NotFound("Orden no encontrada.") from exc
        serializer = serializers.UpdateOrdenMantenimientoSerializer(
            orden, data=request.data, partial=True
        )
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializers.DetailOrdenMantenimientoSerializer(orden).data)


class OrdenMantenimientoCancelarAPIView(APIView):
    """Cancelacion suave de una orden: se marca como CANCE (no se borra
    fisicamente). Se desliga el reporte de falla para que vuelva a estar
    disponible y, si la maquina quedo en MANTE por esta orden, se le
    regresa a ESPER. No se setea fechaCierre a proposito: el trigger
    tg_actualizar_mttr_orden solo se dispara en UPDATE con fechaCierre
    pasando de NULL a valor, y una orden cancelada no es una reparacion."""

    @transaction.atomic
    def patch(self, request, folio):
        try:
            orden = models.OrdenMantenimiento.objects.get(pk=folio)
        except models.OrdenMantenimiento.DoesNotExist as exc:
            raise NotFound("Orden no encontrada.") from exc
        if orden.estado_orden_id in ("CERRA", "CANCE"):
            return Response({"detail": "Esta orden ya está cerrada o cancelada."}, status=400)
        # Las herramientas asignadas a la orden regresan a DISPO (disponibles).
        for ho in models.HerraOrden.objects.filter(orden_mantenimiento=orden):
            liberar_herramienta_de_orden(ho.herramienta_id)
        orden.estado_orden_id = "CANCE"
        orden.reporte_falla = None
        orden.save(update_fields=["estado_orden", "reporte_falla"])
        if orden.maquina_id:
            maquina = Maquina.objects.get(pk=orden.maquina_id)
            if maquina.estado_maquina_id == "MANTE":
                cambiar_estado_maquina(orden.maquina_id, "ESPER", "orden_mantenimiento", orden.folio)
        return Response(serializers.DetailOrdenMantenimientoSerializer(orden).data)


# ------------ EXPORTACIONES  -----------------------------------------
def _get_orden_data(folio):
    qs = models.OrdenMantenimiento.objects.select_related(
        "maquina", "trabajador", "tipo_mantenimiento", "estado_orden", "reporte_falla"
    )
    orden = generics.get_object_or_404(qs, pk=folio)
    ser = serializers.DetailOrdenMantenimientoSerializer(orden)
    return ser.data


class ExportarOrdenCSVAPIView(APIView):

    def get(self, request, folio):
        data = _get_orden_data(folio)
        buffer = io.StringIO()
        writer = csv.writer(buffer)

        writer.writerow(["Campo", "Valor"])
        writer.writerow(["Folio", data.get("folio")])
        writer.writerow(["Descripcion", data.get("descripcion") or ""])
        writer.writerow(["Diagnostico", data.get("diagnostico") or ""])
        writer.writerow(["Notas", data.get("notas") or ""])
        writer.writerow(["Maquina", data.get("maquina_nombre") or ""])
        writer.writerow(["Trabajador", data.get("trabajador_nombre") or ""])
        writer.writerow(["Tipo mantenimiento", data.get("tipo_mantenimiento_nombre") or ""])
        writer.writerow(["Estado", data.get("estado_orden_nombre") or ""])
        writer.writerow(["Fecha creacion", data.get("fechacreacion") or ""])
        writer.writerow(["Hora creacion", data.get("horacreacion") or ""])
        writer.writerow(["Fecha programada", data.get("fechaprogramada") or ""])
        writer.writerow(["Fecha cierre", data.get("fechacierre") or ""])
        writer.writerow(["Hora cierre", data.get("horacierre") or ""])
        writer.writerow(["Horas intervenidas", data.get("horasintervenidas") or ""])
        writer.writerow(["Reporte falla", data.get("reporte_falla_asunto") or ""])

        response = HttpResponse(buffer.getvalue(), content_type="text/csv; charset=utf-8")
        response["Content-Disposition"] = f'attachment; filename="orden_mantenimiento_{folio}.csv"'
        return response


class ExportarOrdenXLSXAPIView(APIView):

    def get(self, request, folio):
        data = _get_orden_data(folio)
        wb = Workbook()
        ws = wb.active
        ws.title = "Orden"

        header_font = Font(bold=True)
        labels = [
            "Folio", "Descripcion", "Diagnostico", "Notas", "Maquina",
            "Trabajador", "Tipo mantenimiento", "Estado", "Fecha creacion",
            "Hora creacion", "Fecha programada", "Fecha cierre",
            "Hora cierre", "Horas intervenidas", "Reporte falla",
        ]
        values = [
            data.get("folio"), data.get("descripcion") or "",
            data.get("diagnostico") or "", data.get("notas") or "",
            data.get("maquina_nombre") or "", data.get("trabajador_nombre") or "",
            data.get("tipo_mantenimiento_nombre") or "",
            data.get("estado_orden_nombre") or "",
            data.get("fechacreacion") or "", data.get("horacreacion") or "",
            data.get("fechaprogramada") or "", data.get("fechacierre") or "",
            data.get("horacierre") or "", data.get("horasintervenidas") or "",
            data.get("reporte_falla_asunto") or "",
        ]
        for row_idx, (label, value) in enumerate(zip(labels, values), 1):
            ws.cell(row=row_idx, column=1, value=label).font = header_font
            ws.cell(row=row_idx, column=2, value=str(value))

        ws.column_dimensions["A"].width = 22
        ws.column_dimensions["B"].width = 60

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        response = HttpResponse(
            buffer.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="orden_mantenimiento_{folio}.xlsx"'
        return response


LOGO_PATH = os.path.join(os.path.dirname(__file__), "assets", "logo.png")
PAGE_W, PAGE_H = letter

BRAND_DARK = colors.HexColor("#12261e")
BRAND_ACCENT = colors.HexColor("#0369a1")
BRAND_DANGER = colors.HexColor("#dc2626")
BRAND_MUTED = colors.HexColor("#6b7280")
BRAND_LINE = colors.HexColor("#e5e7eb")

ESTADO_STYLES = {
    "SOLIC": (colors.HexColor("#e2e8f0"), colors.HexColor("#334155")),
    "PENDI": (colors.HexColor("#e2e8f0"), colors.HexColor("#334155")),
    "APROB": (colors.HexColor("#dbeafe"), colors.HexColor("#1d4ed8")),
    "PROGR": (colors.HexColor("#dbeafe"), colors.HexColor("#1d4ed8")),
    "ENPRO": (colors.HexColor("#fef3c7"), colors.HexColor("#b45309")),
    "ESESP": (colors.HexColor("#fef3c7"), colors.HexColor("#b45309")),
    "EJECU": (colors.HexColor("#dcfce7"), colors.HexColor("#15803d")),
    "CERRA": (colors.HexColor("#dcfce7"), colors.HexColor("#15803d")),
    "CANCE": (colors.HexColor("#fee2e2"), colors.HexColor("#b91c1c")),
}
ESTADO_DEFAULT = (colors.HexColor("#e5e7eb"), colors.HexColor("#374151"))

_styles = getSampleStyleSheet()
_style_label = ParagraphStyle("ordenLabel", parent=_styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=7.5, textColor=BRAND_MUTED, leading=10)
_style_value = ParagraphStyle("ordenValue", parent=_styles["Normal"], fontName="Helvetica",
                               fontSize=9.5, textColor=BRAND_DARK, leading=13)
_style_section_head = ParagraphStyle("ordenSectionHead", parent=_styles["Normal"], fontName="Helvetica-Bold",
                                      fontSize=9.5, textColor=BRAND_DARK, leading=12)
_style_body = ParagraphStyle("ordenBody", parent=_styles["Normal"], fontName="Helvetica",
                              fontSize=9.5, textColor=colors.HexColor("#1f2937"), leading=14)
_style_kicker = ParagraphStyle("ordenKicker", parent=_styles["Normal"], fontName="Helvetica-Bold",
                                fontSize=8, textColor=BRAND_ACCENT, spaceAfter=6)
_style_kicker_falla = ParagraphStyle("fallaKicker", parent=_styles["Normal"], fontName="Helvetica-Bold",
                                      fontSize=8, textColor=BRAND_DANGER, spaceAfter=6)


def _pdf_header_footer(data):
    """Dibuja encabezado y pie en todas las páginas del documento."""
    def _draw(canvas, doc):
        canvas.saveState()
        margin = 20 * mm

        logo_w = 0
        if os.path.isfile(LOGO_PATH):
            logo_w = 14 * mm
            try:
                canvas.drawImage(
                    LOGO_PATH, margin, PAGE_H - margin - logo_w + 6 * mm,
                    width=logo_w, height=logo_w,
                    preserveAspectRatio=True, mask="auto",
                )
            except Exception:
                logo_w = 0

        text_x = margin + (logo_w + 4 * mm if logo_w else 0)
        canvas.setFillColor(BRAND_DARK)
        canvas.setFont("Helvetica-Bold", 14)
        canvas.drawString(text_x, PAGE_H - margin, "OperaCore")
        canvas.setFont("Helvetica", 8)
        canvas.setFillColor(BRAND_MUTED)
        canvas.drawString(text_x, PAGE_H - margin - 10, "Sistema de mantenimiento industrial")

        canvas.setFont("Helvetica-Bold", 12)
        canvas.setFillColor(BRAND_DARK)
        canvas.drawRightString(PAGE_W - margin, PAGE_H - margin, "Orden " + str(data.get("folio") or ""))

        estado_codigo = data.get("estado_orden") or ""
        estado_nombre = (data.get("estado_orden_nombre") or estado_codigo or "—").upper()
        bg, fg = ESTADO_STYLES.get(estado_codigo, ESTADO_DEFAULT)
        badge_w = 8 * mm + 1.7 * mm * len(estado_nombre)
        badge_h = 5.5 * mm
        badge_x = PAGE_W - margin - badge_w
        badge_y = PAGE_H - margin - 10 - badge_h
        canvas.setFillColor(bg)
        canvas.roundRect(badge_x, badge_y, badge_w, badge_h, 2.5, stroke=0, fill=1)
        canvas.setFillColor(fg)
        canvas.setFont("Helvetica-Bold", 8)
        canvas.drawCentredString(badge_x + badge_w / 2, badge_y + 1.7 * mm, estado_nombre)

        canvas.setStrokeColor(BRAND_ACCENT)
        canvas.setLineWidth(1.3)
        line_y = PAGE_H - margin - 20 * mm
        canvas.line(margin, line_y, PAGE_W - margin, line_y)

        canvas.setStrokeColor(BRAND_LINE)
        canvas.setLineWidth(0.6)
        canvas.line(margin, margin - 4 * mm, PAGE_W - margin, margin - 4 * mm)
        canvas.setFont("Helvetica", 7.5)
        canvas.setFillColor(BRAND_MUTED)
        generado = timezone.now().strftime("%d/%m/%Y %H:%M")
        canvas.drawString(margin, margin - 10 * mm, "OperaCore · Documento generado automáticamente el " + generado)
        canvas.drawRightString(PAGE_W - margin, margin - 10 * mm, "Página " + str(doc.page))

        canvas.restoreState()
    return _draw


def _pdf_campo(label, valor):
    return [Paragraph(label.upper(), _style_label), Paragraph(str(valor or "—"), _style_value)]


def _pdf_ficha_tecnica(data):
    filas = [
        _pdf_campo("Máquina", data.get("maquina_nombre")) +
        _pdf_campo("Trabajador asignado", data.get("trabajador_nombre") or "Sin asignar"),

        _pdf_campo("Tipo de mantenimiento", data.get("tipo_mantenimiento_nombre")) +
        _pdf_campo("Horas intervenidas", str(data.get("horasintervenidas") or "—")),

        _pdf_campo("Fecha de creación", (data.get("fechacreacion") or "—") + " " + (data.get("horacreacion") or "")) +
        _pdf_campo("Fecha programada", data.get("fechaprogramada")),

        _pdf_campo("Fecha de cierre", (data.get("fechacierre") or "—") + " " + (data.get("horacierre") or "")) +
        _pdf_campo("Reporte de falla asociado", data.get("reporte_falla_asunto") or "—"),
    ]
    tabla = Table(filas, colWidths=[30 * mm, 54 * mm, 30 * mm, 54 * mm])
    tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, BRAND_LINE),
    ]))
    return tabla


def _pdf_ficha_falla(data):
    filas = [
        _pdf_campo("Severidad", data.get("tipo_severidad_nombre")) +
        _pdf_campo("Estado del reporte", data.get("estado_reporte_nombre")),

        _pdf_campo("Máquina", data.get("maquina_nombre")) +
        _pdf_campo("Trabajador", data.get("trabajador_nombre")),

        _pdf_campo("Fecha de creación", (data.get("fechaCreacion") or "—") + " " + (data.get("horaCreacion") or "")) +
        _pdf_campo("Fecha de resolución", data.get("fechaResolucion") or "—"),

        _pdf_campo("Tiempo de paro (hrs)", str(data.get("tiempoParo") or "—")) +
        _pdf_campo("Tipos de falla", ", ".join(f["nombre"] for f in data.get("fallas_asociadas", [])) or "—"),
    ]
    tabla = Table(filas, colWidths=[30 * mm, 54 * mm, 30 * mm, 54 * mm])
    tabla.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "TOP"),
        ("TOPPADDING", (0, 0), (-1, -1), 6),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 10),
        ("LEFTPADDING", (0, 0), (-1, -1), 0),
        ("LINEBELOW", (0, 0), (-1, -2), 0.5, BRAND_LINE),
    ]))
    return tabla


def _pdf_seccion(titulo, contenido, color=BRAND_ACCENT):
    """Bloque con acento para descripción, diagnóstico y notas."""
    if not contenido:
        return []
    barra = Table([[""]], colWidths=[3], rowHeights=[16])
    barra.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, -1), color)]))
    cuerpo = Table([[barra, Paragraph(titulo.upper(), _style_section_head)]], colWidths=[3, 465])
    cuerpo.setStyle(TableStyle([
        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING", (0, 0), (0, 0), 0),
        ("RIGHTPADDING", (0, 0), (0, 0), 0),
        ("LEFTPADDING", (1, 0), (1, 0), 8),
        ("TOPPADDING", (0, 0), (-1, -1), 0),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 0),
    ]))
    texto = str(contenido).replace("\n", "<br/>")
    return [cuerpo, Spacer(1, 4), Paragraph(texto, _style_body), Spacer(1, 14)]


class ExportarOrdenPDFAPIView(APIView):

    def get(self, request, folio):
        data = _get_orden_data(folio)
        incluir_falla = request.GET.get("incluir_falla") == "1" and bool(data.get("reporte_falla"))

        buffer = io.BytesIO()
        doc = SimpleDocTemplate(
            buffer, pagesize=letter,
            topMargin=34 * mm, bottomMargin=26 * mm,
            leftMargin=20 * mm, rightMargin=20 * mm,
            title="Orden de mantenimiento {}".format(data.get("folio") or ""),
        )

        elements = [
            Paragraph("FICHA TÉCNICA", _style_kicker),
            _pdf_ficha_tecnica(data),
            Spacer(1, 10),
        ]

        for titulo, contenido in [
            ("Descripción", data.get("descripcion")),
            ("Diagnóstico", data.get("diagnostico")),
            ("Notas", data.get("notas")),
        ]:
            elements.extend(_pdf_seccion(titulo, contenido))

        filename_extra = ""
        if incluir_falla:
            falla_data = _get_reporte_data(data["reporte_falla"])
            filename_extra = "_con_falla"

            elements.append(PageBreak())
            elements.append(Paragraph("REPORTE DE FALLA ASOCIADO", _style_kicker_falla))
            elements.append(Paragraph(
                "#{} · {}".format(falla_data.get("numeroRegistro"), falla_data.get("asunto") or ""),
                ParagraphStyle("fallaTitulo", parent=_styles["Normal"], fontName="Helvetica-Bold",
                               fontSize=13, textColor=BRAND_DARK, spaceAfter=10),
            ))
            elements.append(_pdf_ficha_falla(falla_data))
            elements.append(Spacer(1, 10))
            for titulo, contenido in [
                ("Descripción", falla_data.get("descripcion")),
                ("Causa raíz", falla_data.get("causaRaiz")),
            ]:
                elements.extend(_pdf_seccion(titulo, contenido, color=BRAND_DANGER))

        draw = _pdf_header_footer(data)
        doc.build(elements, onFirstPage=draw, onLaterPages=draw)

        buffer.seek(0)
        response = HttpResponse(buffer.getvalue(), content_type="application/pdf")
        response["Content-Disposition"] = "attachment; filename=\"orden_mantenimiento_{}{}.pdf\"".format(folio, filename_extra)
        return response